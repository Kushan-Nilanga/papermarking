from openpyxl import load_workbook, Workbook
import json

# takes in the comment bank file and generate a json dump
def generate_commentbank(filename: str):
    print("\nGenerating comment bank")
    print("loading Comments sheet")
    # load excel file
    worksheet = load_workbook(
        filename=filename)['Comments']

    # data collection
    list_to_json = []

    print("generating comment list")
    # generate a list to turn into json
    for value in worksheet.iter_cols(min_row=1, max_row=3, min_col=3, max_col=80, values_only=True):
        comment = {}
        comment["question"] = value[0]
        comment["correct"] = value[1]
        comment["incorrect"] = value[2]
        list_to_json.append(comment)

    print("writing comment list to bank.json file")
    # writing data to the data.json
    with open('bank.json', 'w') as f:
        f.writelines(json.dumps(list_to_json))


def generate_comments(filename: str):
    print("\nGenerating comments")

    print("loading data bank")
    with open("bank.json") as f:
        bank = json.load(f)

    print("loading students")

    worksheet = load_workbook(
        filename=filename)['Marks']

    wb = Workbook()
    ws = wb.create_sheet("Comments")
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_col=80, values_only=True)):
        ws["A"+str(idx+1)] = row[0]
        ws["B"+str(idx+1)] = row[1]
        ws["C"+str(idx+1)] = row[2]

        comment = ""
        for i in range(3, len(row)):
            if(row[i] == None):
                continue
            placeholder = bank[i -
                               3]['incorrect'] if(row[i] <= 1)else bank[i-3]['correct']
            if(placeholder != None and placeholder != "null"):
                comment += placeholder
            comment += "\n"
        ws["D"+str(idx+1)] = comment
    wb.save("comments.xlsx")


def generate_marks(filename: str):
    print("hello")
